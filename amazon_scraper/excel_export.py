from __future__ import annotations

import asyncio
from io import BytesIO

import httpx
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from .models import BatchResult, CompetitorExportRequest, TitleExportRequest


HEADERS = [
    "图片", "图片 URL", "输入 ASIN", "父体 ASIN", "疑似主销", "子体 ASIN", "标题",
    "颜色", "尺寸", "当前价格", "Typical price", "折扣", "月销量信号",
    "月销量估算", "评分", "评分数", "库存", "数据完整度", "判断依据",
]


async def _download_image(
    client: httpx.AsyncClient,
    url: str | None,
    semaphore: asyncio.Semaphore,
) -> BytesIO | None:
    if not url:
        return None
    try:
        async with semaphore:
            response = await client.get(url, timeout=20)
            response.raise_for_status()
        source = PILImage.open(BytesIO(response.content))
        source.thumbnail((82, 82))
        if source.mode not in ("RGB", "RGBA"):
            source = source.convert("RGBA")
        output = BytesIO()
        source.save(output, format="PNG")
        output.seek(0)
        return output
    except Exception:
        return None


async def build_products_xlsx(batch: BatchResult) -> BytesIO:
    rows: list[tuple] = []
    image_urls: list[str | None] = []
    for item in batch.items:
        if not item.success or not item.result:
            rows.append(("", "", item.requested_asin, "", "", "", f"采集失败：{item.error or ''}", "", "", "", "", "", "", "", "", "", "", "失败", ""))
            image_urls.append(None)
            continue
        product = item.result
        for variant in product.variants:
            rows.append((
                "", variant.image or "", item.requested_asin, product.parent_asin or "",
                (
                    "是" if variant.is_suspected_main else "否"
                ) if product.is_parent_request else "",
                variant.asin,
                variant.title or "", variant.color or "", variant.size or "",
                variant.price or "", variant.list_price or "", variant.discount or "",
                variant.recent_sales_signal or "", variant.monthly_sales_estimate,
                variant.rating, variant.rating_count, variant.availability or "",
                "完整" if variant.data_quality == "complete" else "部分",
                variant.main_reason or "",
            ))
            image_urls.append(variant.image)

    semaphore = asyncio.Semaphore(8)
    async with httpx.AsyncClient(
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
            "Referer": "https://www.amazon.com/",
        },
        follow_redirects=True,
    ) as client:
        images = await asyncio.gather(*[
            _download_image(client, url, semaphore) for url in image_urls
        ])

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品子体"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="173F35")
    main_fill = PatternFill("solid", fgColor="E8F4EE")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 26
    sheet.freeze_panes = "B2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{sheet.max_row}"

    image_streams: list[BytesIO] = []
    for index, image_stream in enumerate(images, start=2):
        sheet.row_dimensions[index].height = 68
        image_url_cell = sheet.cell(index, 2)
        if image_url_cell.value:
            image_url_cell.hyperlink = str(image_url_cell.value)
            image_url_cell.style = "Hyperlink"
        if sheet.cell(index, 5).value == "是":
            for cell in sheet[index]:
                cell.fill = main_fill
        for cell in sheet[index]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        if image_stream:
            image_streams.append(image_stream)
            image = ExcelImage(image_stream)
            image.width = 82
            image.height = 82
            sheet.add_image(image, f"A{index}")

    widths = [14, 42, 15, 15, 11, 15, 54, 18, 16, 14, 16, 12, 24, 14, 10, 12, 22, 12, 38]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


async def build_competitors_xlsx(request: CompetitorExportRequest) -> BytesIO:
    headers = [
        "图片", "图片 URL", "已选", "本品 ASIN", "竞品 ASIN", "竞品父体 ASIN",
        "品牌", "尺寸", "标题", "价格", "评分", "评分数", "月销量信号",
        "月销量估算", "总匹配分", "标题词分", "属性分", "视觉分", "视觉依据",
        "市场分", "匹配依据", "商品链接",
    ]
    semaphore = asyncio.Semaphore(8)
    async with httpx.AsyncClient(
        headers={"User-Agent": "Mozilla/5.0", "Referer": "https://www.amazon.com/"},
        follow_redirects=True,
    ) as client:
        images = await asyncio.gather(*[
            _download_image(client, item.image, semaphore) for item in request.items
        ])

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "竞品候选"
    sheet.append(headers)
    for item in request.items:
        sheet.append([
            "", item.image or "", "是" if item.selected else "否", request.target_asin or "",
            item.asin, item.parent_asin or "", item.brand or "", item.size or "",
            item.title, item.price or "", item.rating, item.rating_count,
            item.recent_sales_signal or "", item.monthly_sales_estimate,
            item.overall_similarity, item.text_similarity, item.attribute_similarity,
            item.image_similarity, item.visual_reason or "", item.market_similarity,
            "；".join(item.match_reasons), item.url,
        ])

    header_fill = PatternFill("solid", fgColor="173F35")
    selected_fill = PatternFill("solid", fgColor="E8F4EE")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 26
    sheet.freeze_panes = "I2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"

    image_streams: list[BytesIO] = []
    for row_index, image_stream in enumerate(images, start=2):
        sheet.row_dimensions[row_index].height = 68
        for column in (2, 22):
            cell = sheet.cell(row_index, column)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
        if sheet.cell(row_index, 3).value == "是":
            for cell in sheet[row_index]:
                cell.fill = selected_fill
        for cell in sheet[row_index]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        if image_stream:
            image_streams.append(image_stream)
            image = ExcelImage(image_stream)
            image.width = 82
            image.height = 82
            sheet.add_image(image, f"A{row_index}")

    widths = [14, 38, 8, 15, 15, 16, 16, 14, 58, 12, 9, 11, 24, 13, 11, 11, 10, 10, 50, 10, 48, 38]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_titles_xlsx(request: TitleExportRequest) -> BytesIO:
    """Export one auditable title matrix with sizes as the requested column headers."""
    strategy_names = {"traffic": "流量优先", "click": "点击吸引", "balanced": "均衡转化"}
    sizes = list(dict.fromkeys(item.size.strip() or "尺寸未识别" for item in request.items))
    by_size_strategy = {(item.size.strip() or "尺寸未识别", item.strategy): item for item in request.items}

    workbook = Workbook()
    matrix = workbook.active
    matrix.title = "按尺寸标题"
    matrix.append(["标题字段", *sizes])
    row_specs = [
        ("主标题（流量优先）", "traffic", "main_title"),
        ("Item Highlight（流量优先）", "traffic", "highlight_item"),
        ("完整标题（流量优先）", "traffic", "full_title"),
        ("主标题（点击吸引）", "click", "main_title"),
        ("Item Highlight（点击吸引）", "click", "highlight_item"),
        ("完整标题（点击吸引）", "click", "full_title"),
        ("主标题（均衡转化）", "balanced", "main_title"),
        ("Item Highlight（均衡转化）", "balanced", "highlight_item"),
        ("完整标题（均衡转化）", "balanced", "full_title"),
    ]
    for label, strategy, field in row_specs:
        matrix.append([
            label,
            *[
                getattr(by_size_strategy.get((size, strategy)), field, "") or ""
                for size in sizes
            ],
        ])

    details = workbook.create_sheet("标题明细")
    details.append([
        "父体 ASIN", "主推子体 ASIN", "尺寸", "策略", "主标题", "Item Highlight",
        "完整标题", "评分", "覆盖关键词", "风险提示",
    ])
    for item in request.items:
        details.append([
            request.parent_asin or "", request.main_child_asin or "", item.size,
            strategy_names[item.strategy], item.main_title, item.highlight_item or "",
            item.full_title, item.score, "；".join(item.keywords_used),
            "；".join(item.warnings),
        ])

    header_fill = PatternFill("solid", fgColor="173F35")
    label_fill = PatternFill("solid", fgColor="E8F4EE")
    for sheet in (matrix, details):
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.freeze_panes = "B2"
        sheet.row_dimensions[1].height = 28
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_index in range(2, matrix.max_row + 1):
        matrix.cell(row_index, 1).fill = label_fill
        matrix.cell(row_index, 1).font = Font(bold=True, color="173F35")
        matrix.row_dimensions[row_index].height = 60
    matrix.column_dimensions["A"].width = 30
    for column in range(2, matrix.max_column + 1):
        matrix.column_dimensions[get_column_letter(column)].width = 52
    detail_widths = [15, 15, 16, 12, 54, 54, 70, 9, 42, 44]
    for index, width in enumerate(detail_widths, start=1):
        details.column_dimensions[get_column_letter(index)].width = width
    details.auto_filter.ref = f"A1:J{details.max_row}"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

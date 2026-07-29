from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .models import KeywordEntry, KeywordFileSummary

KEYWORD_ALIASES = ("关键词", "搜索词", "keyword", "search term", "query")
VOLUME_ALIASES = ("搜索量", "月搜索量", "search volume", "流量", "预估搜索量")
MONTH_ALIASES = ("月份", "month", "日期", "date")


def _clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _matches(value: str, aliases: tuple[str, ...]) -> bool:
    lowered = value.lower()
    return any(alias in lowered for alias in aliases)


def _numeric(value: object) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"[\d,.]+", _clean(value))
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", ""))
    except ValueError:
        return None


def _date_key(value: str, index: int) -> tuple[int, int, int]:
    numbers = [int(number) for number in re.findall(r"\d+", value)]
    if len(numbers) >= 2:
        year, month = numbers[0], numbers[1]
        if year < 100:
            year += 2000
        if 1 <= month <= 12:
            return year, month, index
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m", "%Y/%m"):
        try:
            parsed = datetime.strptime(value.strip(), pattern)
            return parsed.year, parsed.month, index
        except ValueError:
            pass
    return 0, 0, index


def inspect_keyword_file(filename: str, content: bytes) -> KeywordFileSummary:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xls":
        raise ValueError("旧版 .xls 暂不支持，请另存为 .xlsx 后上传。")
    if suffix not in {".xlsx", ".csv"}:
        raise ValueError("ABA 综合词库请上传 .xlsx 或 .csv 文件。")
    if not content:
        raise ValueError("上传的词库为空。")

    sheet_name = "CSV"
    if suffix == ".xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        sheet_name = sheet.title
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    else:
        text = content.decode("utf-8-sig", errors="replace")
        rows = [row for row in csv.reader(io.StringIO(text))]

    header_index = -1
    keyword_index = -1
    volume_indexes: list[int] = []
    month_indexes: list[int] = []
    headers: list[str] = []
    for row_index, row in enumerate(rows[:30]):
        current = [_clean(value) for value in row]
        possible_keyword = next(
            (index for index, value in enumerate(current) if _matches(value, KEYWORD_ALIASES)),
            -1,
        )
        if possible_keyword < 0:
            continue
        header_index = row_index
        keyword_index = possible_keyword
        headers = current
        volume_indexes = [
            index for index, value in enumerate(current) if _matches(value, VOLUME_ALIASES)
        ]
        month_indexes = [
            index for index, value in enumerate(current) if _matches(value, MONTH_ALIASES)
        ]
        break

    if header_index < 0:
        return KeywordFileSummary(
            filename=filename,
            sheet=sheet_name,
            valid=False,
            warnings=["前 30 行未找到“关键词 / 搜索词 / Keyword”列。"],
        )

    # Monthly search-volume columns are snapshots. Use the latest dated column;
    # when dates are absent, the right-most volume column is treated as latest.
    latest_volume_index = max(
        volume_indexes,
        key=lambda index: _date_key(headers[index], index),
        default=-1,
    )
    latest_label = headers[latest_volume_index] if len(volume_indexes) > 1 and latest_volume_index >= 0 else None
    keyword_map: dict[str, KeywordEntry] = {}
    keyword_dates: dict[str, tuple[int, int, int]] = {}
    for row in rows[header_index + 1:]:
        value = _clean(row[keyword_index] if keyword_index < len(row) else "")
        if not value:
            continue
        volume = (
            _numeric(row[latest_volume_index])
            if latest_volume_index >= 0 and latest_volume_index < len(row)
            else None
        )
        month = latest_label or next(
            (_clean(row[index]) for index in month_indexes if index < len(row) and _clean(row[index])),
            None,
        )
        key = value.lower()
        current = keyword_map.get(key)
        row_date = _date_key(month or "", 0)
        current_date = keyword_dates.get(key, (0, 0, 0))
        if current is None or row_date > current_date or (
            row_date == current_date and (volume or 0) > (current.volume or 0)
        ):
            keyword_map[key] = KeywordEntry(term=value, volume=volume, month=month)
            keyword_dates[key] = row_date

    keyword_entries = sorted(
        keyword_map.values(),
        key=lambda item: (-(item.volume or 0), item.term.lower()),
    )
    keyword_entries = [
        item.model_copy(update={"rank": index})
        for index, item in enumerate(keyword_entries, start=1)
    ]

    warnings: list[str] = []
    if not volume_indexes:
        warnings.append("未识别到搜索量列；仍可进入下一步，但无法按流量权重选词。")
    if not keyword_entries:
        warnings.append("关键词列下没有有效数据。")
    return KeywordFileSummary(
        filename=filename,
        sheet=sheet_name,
        valid=bool(keyword_entries),
        rows=len(keyword_entries),
        keyword_column=headers[keyword_index],
        volume_columns=[headers[index] for index in volume_indexes],
        month_columns=[headers[index] for index in month_indexes],
        preview=[item.term for item in keyword_entries[:5]],
        warnings=warnings,
        keywords=keyword_entries[:10_000],
    )


def inspect_negative_file(filename: str, content: bytes) -> dict[str, object]:
    suffix = Path(filename).suffix.lower()
    if suffix not in {".xlsx", ".csv", ".txt"}:
        raise ValueError("否词词库请上传 .xlsx、.csv 或 .txt 文件。")
    if not content:
        raise ValueError("上传的否词词库为空。")
    values: list[str] = []
    if suffix == ".xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            values.extend(_clean(value) for value in row if _clean(value))
    else:
        text = content.decode("utf-8-sig", errors="replace")
        for row in csv.reader(io.StringIO(text)):
            values.extend(_clean(value) for value in row if _clean(value))
    header_aliases = {*KEYWORD_ALIASES, "否词", "禁用词", "negative keyword", "negative term"}
    terms = []
    seen: set[str] = set()
    for value in values:
        normalized = value.lower()
        if normalized in header_aliases or normalized in seen or len(value) > 100:
            continue
        seen.add(normalized)
        terms.append(value)
    return {"filename": filename, "valid": bool(terms), "rows": len(terms), "terms": terms[:10_000]}
